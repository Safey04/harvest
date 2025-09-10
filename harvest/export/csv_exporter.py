"""
CSV export functionality for harvest optimization results.
"""

import pandas as pd
import os
from typing import List, Dict, Any
from datetime import datetime
import logging

from ..models import HarvestEntry, RejectionEntry, SummaryEntry, DailySummary

logger = logging.getLogger(__name__)


class CSVExporter:
    """
    Responsible for exporting structured data to CSV files.
    Follows Single Responsibility Principle - only handles CSV export.
    """
    
    def __init__(self, output_dir: str = "output", scenario_name: str = ""):
        """
        Initialize CSV exporter.
        
        Args:
            output_dir: Directory to save CSV files
            scenario_name: Name of the scenario to append to filenames (e.g., "age_expansion")
        """
        self.output_dir = output_dir
        self.scenario_name = scenario_name
        self._ensure_output_dir()
    
    def _ensure_output_dir(self):
        """Ensure output directory exists."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            logger.info(f"Created output directory: {self.output_dir}")
    
    def _get_filename_with_scenario(self, base_filename: str) -> str:
        """
        Get filename with scenario name appended if provided.
        
        Args:
            base_filename: Base filename without extension
            
        Returns:
            Filename with scenario name appended if scenario_name is provided
        """
        if self.scenario_name:
            # Split filename and extension
            name, ext = os.path.splitext(base_filename)
            return f"{name}_{self.scenario_name}{ext}"
        return base_filename
    
    def export_harvest_plan_master(self, harvest_df: pd.DataFrame) -> str:
        """
        Export master harvest plan as professional pivoted Excel file with hierarchical columns.
        Structure: Index (Date, House) | Columns (Farm > Harvest_Type > [Harvest_Stock, Avg_Weight])
        
        Args:
            harvest_df: Combined harvest DataFrame
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_plan_master_pivot.xlsx"))
        
        # Ensure required columns exist
        required_columns = [
            'farm', 'date', 'house', 'age', 'harvest_type', 'harvest_stock',
            'expected_stock', 'avg_weight', 'net_meat', 'expected_mortality',
            'expected_mortality_rate'
        ]
        
        export_df = harvest_df[required_columns].copy()
        
        # Export raw data to CSV for reference
        csv_filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_plan_master.csv"))
        export_df.to_csv(csv_filename, index=True)
        logger.info(f"Exported raw harvest plan data to {csv_filename}")
        
        # Debug: Check unique dates before processing
        logger.info(f"Unique dates in data: {export_df['date'].unique()}")
        logger.info(f"Data shape: {export_df.shape}")
        logger.info(f"Sample of first few rows:")
        logger.info(export_df[['date', 'house', 'farm', 'harvest_type']].head(10))
        
        # Ensure date is properly formatted
        export_df['date'] = pd.to_datetime(export_df['date'])
        
        # Debug: Check after date conversion
        logger.info(f"Unique dates after conversion: {export_df['date'].unique()}")
        logger.info(f"Date range: {export_df['date'].min()} to {export_df['date'].max()}")
        
        # Create the pivot table with the specified structure
        # Index: date, house
        # Columns: farm, harvest_type  
        # Values: harvest_stock and avg_weight
        pivot_df = export_df.pivot_table(
            index=['date', 'house'],
            columns=['farm', 'harvest_type'],
            values=['harvest_stock', 'avg_weight'],
            aggfunc={
                'harvest_stock': 'sum',
                'avg_weight': 'mean'
            },
            fill_value=None
        )
        
        # Debug: Check pivot table structure
        logger.info(f"Pivot table shape: {pivot_df.shape}")
        logger.info(f"Pivot table index levels: {pivot_df.index.names}")
        logger.info(f"Pivot table columns levels: {pivot_df.columns.names}")
        logger.info(f"Sample index values: {pivot_df.index[:5]}")
        logger.info(f"All unique dates in pivot index: {sorted(pivot_df.index.get_level_values('date').unique())}")
        logger.info(f"All unique houses in pivot index: {sorted(pivot_df.index.get_level_values('house').unique())}")
        
        # Reorder columns to group harvest_stock and avg_weight under each farm-harvest_type combination
        # This creates the structure: Farm > Harvest_Type > [Harvest_Stock, Avg_Weight]
        new_columns = []
        
        # Get unique farm and harvest_type combinations from the original data
        combinations = export_df[['farm', 'harvest_type']].drop_duplicates().sort_values(['farm', 'harvest_type'])
        
        for _, row in combinations.iterrows():
            farm = row['farm']
            harvest_type = row['harvest_type']
            
            # Add harvest_stock column for this combination
            if ('harvest_stock', farm, harvest_type) in pivot_df.columns:
                new_columns.append(('harvest_stock', farm, harvest_type))
            
            # Add avg_weight column for this combination
            if ('avg_weight', farm, harvest_type) in pivot_df.columns:
                new_columns.append(('avg_weight', farm, harvest_type))
        
        # Reorder the pivot table columns
        pivot_df = pivot_df.reindex(columns=new_columns)
        
        # Sort index by date, then house
        pivot_df = pivot_df.sort_index()
        
        # Debug: Check final pivot table
        logger.info(f"Final pivot table shape: {pivot_df.shape}")
        logger.info(f"Final index levels: {pivot_df.index.names}")
        logger.info(f"Final sample index: {pivot_df.index[:5]}")
        
        # Format dates for better Excel display - but preserve the multi-level index structure
        # Convert the index to a DataFrame, format dates, then back to index
        index_df = pivot_df.index.to_frame()
        
        # Debug: Check what dates we have before formatting
        logger.info(f"Unique dates in index before formatting: {index_df['date'].unique()}")
        logger.info(f"Date column dtype: {index_df['date'].dtype}")
        
        # Ensure dates are properly converted and formatted
        if index_df['date'].dtype == 'object':
            # If dates are already strings, convert to datetime first
            index_df['date'] = pd.to_datetime(index_df['date'])
        
        # Format dates consistently
        index_df['date'] = index_df['date'].dt.strftime('%Y-%m-%d')
        
        # Debug: Check what dates we have after formatting
        logger.info(f"Unique dates in index after formatting: {index_df['date'].unique()}")
        
        # Create new index with formatted dates
        new_index = pd.MultiIndex.from_frame(index_df)
        pivot_df.index = new_index
        
        # Export to Excel with professional formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Reset the index to make date and house regular columns for better Excel handling
            excel_df = pivot_df.reset_index()
            
            # Flatten the MultiIndex columns to make them Excel-friendly
            # Create new column names that combine the levels
            new_columns = []
            for col in excel_df.columns:
                if isinstance(col, tuple):
                    # For MultiIndex columns, combine the levels
                    if col[0] == 'harvest_stock':
                        new_columns.append(f"{col[1]}_{col[2]}_Stock")
                    elif col[0] == 'avg_weight':
                        new_columns.append(f"{col[1]}_{col[2]}_Weight")
                    else:
                        new_columns.append("_".join(str(level) for level in col))
                else:
                    # For regular columns (date, house), keep as is
                    new_columns.append(col)
            
            excel_df.columns = new_columns
            
            # Export the flattened DataFrame
            excel_df.to_excel(writer, sheet_name='Harvest Plan Master', index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Harvest Plan Master']
            
            # Apply professional formatting
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Define styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            data_font = Font(name='Calibri', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Add title row first, before any formatting
            worksheet.insert_rows(1)
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = "Harvest Plan Master Report"
            title_cell.font = Font(name='Calibri', size=14, bold=True)
            worksheet.merge_cells(f'A1:{get_column_letter(worksheet.max_column)}1')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format header rows - now we have a simpler structure with title + column headers
            for row in range(2, 3):  # Row 2 for column headers (after title)
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = border
            
            # Format date and house columns - starting from row 3 due to title + header
            for row in range(3, worksheet.max_row + 1):
                for col in range(1, 3):  # First two columns are date and house
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    if col == 1:  # Date column
                        cell.alignment = Alignment(horizontal='center')
                    else:  # House column
                        cell.alignment = Alignment(horizontal='center')
            
            # Format data cells - starting from row 3 due to title + header
            for row in range(3, worksheet.max_row + 1):
                for col in range(3, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
                    # Format numbers based on column type
                    column_header = worksheet.cell(row=2, column=col).value  # Row 2 is now the header row
                    if column_header and 'weight' in str(column_header).lower():
                        cell.number_format = '0.00'  # 2 decimal places for weights
                    elif column_header and 'stock' in str(column_header).lower():
                        cell.number_format = '#,##0'  # Integer format with commas for stock
            
            # Auto-adjust column widths
            for col_num in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                # Check all rows in this column for content length
                for row_num in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value is not None:
                        try:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                        except:
                            pass
                
                # Set column width with reasonable bounds
                adjusted_width = min(max(max_length + 2, 10), 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes at the intersection of headers and data (after title row)
            worksheet.freeze_panes = 'C3'  # Freeze first 2 columns (date and house) and first 2 rows (title + header)
        
        logger.info(f"Exported professional pivoted master harvest plan to {filename}")
        return filename

    

    def export_harvest_plan_slaughterhouse(self, harvest_df: pd.DataFrame) -> str:
        """
        Export slaughterhouse-specific harvest plan CSV.
        
        Args:
            harvest_df: Harvest DataFrame filtered for slaughterhouse
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_plan_slaughterhouse.csv"))
        
        sh_df = harvest_df[harvest_df['harvest_type'] == 'SH'].copy()
        
        if not sh_df.empty:
            sh_df['date'] = pd.to_datetime(sh_df['date']).dt.strftime('%Y-%m-%d')
            sh_df.to_csv(filename, index=False)
            columns_to_export = ['farm','date','house','age','harvest_type','harvest_stock','expected_stock','avg_weight','net_meat','expected_mortality','expected_mortality_rate']
            sh_df = sh_df[columns_to_export]
            logger.info(f"Exported slaughterhouse harvest plan to {filename}")
        else:
            # Create empty file with headers
            pd.DataFrame(columns=harvest_df.columns).to_csv(filename, index=False)
            logger.warning(f"No slaughterhouse data found, created empty file: {filename}")
        
        return filename
    
    def export_harvest_plan_market(self, harvest_df: pd.DataFrame) -> str:
        """
        Export market-specific harvest plan CSV.
        
        Args:
            harvest_df: Harvest DataFrame filtered for market
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_plan_market.csv"))

        
        
        market_df = harvest_df[harvest_df['harvest_type'] == 'Market'].copy()
        
        if not market_df.empty:
            market_df['date'] = pd.to_datetime(market_df['date']).dt.strftime('%Y-%m-%d')

            #only export these columns
            columns_to_export = ['farm','date','house','age','harvest_type','harvest_stock','expected_stock','avg_weight','net_meat','expected_mortality','expected_mortality_rate','priority','profit_per_bird']
            market_df = market_df[columns_to_export]
            market_df.to_csv(filename, index=False)
            logger.info(f"Exported market harvest plan to {filename}")
        else:
            # Create empty file with headers
            pd.DataFrame(columns=harvest_df.columns).to_csv(filename, index=False)
            logger.warning(f"No market data found, created empty file: {filename}")
        
        return filename
    
    def export_harvest_plan_culls(self, cull_df: pd.DataFrame) -> str:
        """
        Export cull harvest plan CSV.
        
        Args:
            cull_df: DataFrame with cull data
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_plan_culls.csv"))
        
        if not cull_df.empty:
            cull_df['date'] = pd.to_datetime(cull_df['date']).dt.strftime('%Y-%m-%d')
            cull_df = cull_df[['farm', 'date', 'house', 'age', 'harvest_type', 'harvest_stock', 'expected_stock', 'avg_weight', 'net_meat']]
            cull_df.to_csv(filename, index=False)
            logger.info(f"Exported cull harvest plan to {filename}")
        else:
            # Create empty file with headers
            headers = ['farm', 'date', 'house', 'age', 'harvest_type', 'harvest_stock', 
                      'expected_stock', 'avg_weight', 'net_meat']
            pd.DataFrame(columns=headers).to_csv(filename, index=False)
            logger.warning(f"No cull data found, created empty file: {filename}")
        
        return filename
    
    def export_summary_farms(self, harvest_df: pd.DataFrame) -> str:
        """
        Export farm summary CSV.
        
        Args:
            harvest_df: Combined harvest DataFrame
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("summary_farms.csv"))
        
        # Group by farm and harvest type
        farm_summary = harvest_df.groupby(['farm', 'harvest_type']).agg({
            'harvest_stock': 'sum',
            'net_meat': 'sum',
            'house': 'nunique',
            'date': 'nunique'
        }).reset_index()
        
        farm_summary.rename(columns={
            'harvest_stock': 'total_harvest_stock',
            'net_meat': 'total_net_meat',
            'house': 'unique_houses',
            'date': 'harvest_days'
        }, inplace=True)
        
        # Calculate kg_density based on total net meat per farm (across all harvest types)
        # First get total net meat per farm
        farm_totals = harvest_df.groupby(['farm'])['net_meat'].sum().reset_index()
        farm_totals.rename(columns={'net_meat': 'farm_total_net_meat'}, inplace=True)
        
        # Merge with farm_summary
        farm_summary = farm_summary.merge(farm_totals, on=['farm'])
        
        # Calculate kg_density (farm_total_net_meat / farm_area)
        # Farm area = 1680 * 12 houses = 20160
        farm_summary['kg_density'] = farm_summary['farm_total_net_meat'] / 20160
        
        # Drop the helper column
        farm_summary.drop('farm_total_net_meat', axis=1, inplace=True)
        
        farm_summary.to_csv(filename, index=False)
        logger.info(f"Exported farm summary to {filename}")
        
        return filename
    
    def export_summary_houses(self, harvest_df: pd.DataFrame) -> str:
        """
        Export house summary CSV.
        
        Args:
            harvest_df: Combined harvest DataFrame
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("summary_houses.csv"))
        
        # Group by farm, house, and harvest type
        house_summary = harvest_df.groupby(['farm', 'house', 'harvest_type']).agg({
            'harvest_stock': 'sum',
            'net_meat': 'sum',
            'date': 'nunique'
        }).reset_index()
        
        house_summary.rename(columns={
            'harvest_stock': 'total_harvest_stock',
            'net_meat': 'total_net_meat',
            'date': 'harvest_days'
        }, inplace=True)
        
        # Calculate kg_density based on total net meat per house (across all harvest types)
        # First get total net meat per house
        house_totals = harvest_df.groupby(['farm', 'house'])['net_meat'].sum().reset_index()
        house_totals.rename(columns={'net_meat': 'house_total_net_meat'}, inplace=True)
        
        # Merge with house_summary
        house_summary = house_summary.merge(house_totals, on=['farm', 'house'])
        
        # Calculate kg_density (house_total_net_meat / house_area)
        # House area = 1680
        house_summary['kg_density'] = house_summary['house_total_net_meat'] / 1680
        
        # Drop the helper column
        house_summary.drop('house_total_net_meat', axis=1, inplace=True)
        
        house_summary.to_csv(filename, index=False)
        logger.info(f"Exported house summary to {filename}")
        
        return filename
    
    def export_unharvested_stock(self, unharvested_df: pd.DataFrame, ready_df: pd.DataFrame, best_market_plan: pd.DataFrame, best_sh_plan: pd.DataFrame) -> str:
        """
        Export unharvested stock CSV.
        
        Args:
            unharvested_df: DataFrame with unharvested stock
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("unharvested_stock.csv"))

        # Find records in ready_df but not in unharvested_df and not in best_market_plan
        # Get unique farm-house pairs from unharvested_df and best_market_plan
        if not unharvested_df.empty and 'farm' in unharvested_df.columns and 'house' in unharvested_df.columns:
            unharvested_pairs = unharvested_df[['farm', 'house']].drop_duplicates()
        else:
            unharvested_pairs = pd.DataFrame(columns=['farm', 'house'])
            
        if not best_market_plan.empty and 'farm' in best_market_plan.columns and 'house' in best_market_plan.columns:
            market_plan_pairs = best_market_plan[['farm', 'house']].drop_duplicates()
        else:
            market_plan_pairs = pd.DataFrame(columns=['farm', 'house'])
        
        # Combine unharvested and market plan pairs
        excluded_pairs = pd.concat([unharvested_pairs, market_plan_pairs]).drop_duplicates()
        
        # Find records in ready_df but not in excluded pairs (anti-join)
        not_in_both = ready_df.merge(excluded_pairs, on=['farm', 'house'], how='left', indicator=True)
        not_in_both = not_in_both[not_in_both['_merge'] == 'left_only'].drop(columns=['_merge'])
        not_in_both = not_in_both[not_in_both['age'] >= 28]
        
        # Subtract culls and slaughterhouse allocations from expected_stock
        if not not_in_both.empty:
            # Get culls and slaughterhouse allocations for these farm-house combinations
            culls_allocations = best_market_plan[best_market_plan[['farm', 'house']].apply(
                lambda x: x.isin(not_in_both[['farm', 'house']]).all(), axis=1
            )]['harvest_stock'].sum() if not best_market_plan.empty else 0
            
            sh_allocations = best_sh_plan[best_sh_plan[['farm', 'house']].apply(
                lambda x: x.isin(not_in_both[['farm', 'house']]).all(), axis=1
            )]['harvest_stock'].sum() if 'best_sh_plan' in locals() and not best_sh_plan.empty else 0
            
            # Subtract allocations from expected_stock
            not_in_both['expected_stock'] = not_in_both['expected_stock'] - (culls_allocations + sh_allocations)
            not_in_both['expected_stock'] = not_in_both['expected_stock'].clip(lower=0)  
        
        # Add reason column
        not_in_both['unharvested_reason'] = 'Did not meet Minimum Average Weight Constraint'
        


        
        if not unharvested_df.empty and 'farm' in unharvested_df.columns and 'house' in unharvested_df.columns:
            export_df = unharvested_df.copy()
            export_df['date'] = pd.to_datetime(export_df['date']).dt.strftime('%Y-%m-%d')
            export_df = export_df.drop_duplicates(subset=['farm', 'house', 'date', 'age'])
            export_df['unharvested_reason'] = 'Exceeded Capacity & not optimal Average Weight'
            export_df = pd.concat([export_df, not_in_both])
            columns_to_export =  ["farm","date","house","age","expected_mortality","expected_stock","expected_mortality_rate","avg_weight","unharvested_reason"]
            export_df = export_df[columns_to_export]
            export_df = export_df[export_df['expected_stock'] > 1550]
            export_df.to_csv(filename, index=False)
            logger.info(f"Exported unharvested stock to {filename}")
        elif not not_in_both.empty: 
            not_in_both.to_csv(filename, index=False)
            logger.info(f"Exported unharvested stock to {filename}")
            
        else:
            # Create empty file with headers
            headers = ['farm', 'house', 'date', 'age', 'expected_stock', 'avg_weight','unharvested_reason']
            pd.DataFrame(columns=headers).to_csv(filename, index=False)
            logger.info(f"No unharvested stock, created empty file: {filename}")
        
        return filename
    
    def export_rejection_reasons(self, rejection_df: pd.DataFrame) -> str:
        """
        Export rejection reasons CSV.
        
        Args:
            rejection_df: DataFrame with rejection data
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("rejection_dates.csv"))
        
        if not rejection_df.empty:
            export_df = rejection_df.copy()
            if 'date' in export_df.columns:
                export_df['date'] = pd.to_datetime(export_df['date']).dt.strftime('%Y-%m-%d')
            export_df.to_csv(filename, index=False)
            logger.info(f"Exported rejection reasons to {filename}")
        else:
            # Create empty file with headers
            headers = ['farm', 'house', 'date', 'age', 'expected_stock', 'avg_weight', 'rejection_reason']
            pd.DataFrame(columns=headers).to_csv(filename, index=False)
            logger.info(f"No rejection data, created empty file: {filename}")
        
        return filename

    def export_harvest_progression(self, harvest_df_progress: pd.DataFrame) -> str:
        """
        Export harvest progression CSV.
        
        Args:
            harvest_df: Combined harvest DataFrame
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("harvest_progression.csv"))
        
        harvest_df_progress.to_csv(filename, index=False)
        logger.info(f"Exported harvest progression to {filename}")
        
        return filename


    
    def export_daily_harvest_summary(self, harvest_df: pd.DataFrame) -> str:
        """
        Export daily harvest summary CSV.
        
        Args:
            harvest_df: Combined harvest DataFrame
            
        Returns:
            Path to exported file
        """
        filename = os.path.join(self.output_dir, self._get_filename_with_scenario("daily_harvest_summary.csv"))
        
        # Group by date and harvest type
        daily_summary = harvest_df.groupby(['date', 'harvest_type']).agg({
            'harvest_stock': 'sum',
            'net_meat': 'sum',
            'house': 'nunique'
        }).reset_index()
        
        daily_summary.rename(columns={
            'harvest_stock': 'total_stock',
            'net_meat': 'total_net_meat',
            'house': 'houses_harvested'
        }, inplace=True)
        
        daily_summary['date'] = pd.to_datetime(daily_summary['date']).dt.strftime('%Y-%m-%d')
        daily_summary.to_csv(filename, index=False)
        logger.info(f"Exported daily harvest summary to {filename}")
        
        return filename


class HarvestPlanExporter:
    """
    High-level exporter that coordinates all CSV exports.
    """
    
    def __init__(self, output_dir: str = "output", scenario_name: str = ""):
        """
        Initialize harvest plan exporter.
        
        Args:
            output_dir: Directory to save CSV files
            scenario_name: Name of the scenario to append to filenames (e.g., "age_expansion")
        """
        self.csv_exporter = CSVExporter(output_dir, scenario_name)
    
    def generate_harvest_plan_csvs(
        self, 
        harvest_df: pd.DataFrame,
        unharvested_df: pd.DataFrame = None,
        rejection_df: pd.DataFrame = None,
        ready_df: pd.DataFrame = None,
        best_sh_plan: pd.DataFrame = None,
        best_market_plan: pd.DataFrame = None,
        cull_df: pd.DataFrame = None,
        harvest_df_progress: pd.DataFrame = None

    ) -> Dict[str, str]:
        """
        Generate all required CSV files for harvest plan.
        
        Args:
            harvest_df: Combined harvest DataFrame
            unharvested_df: DataFrame with unharvested stock
            rejection_df: DataFrame with rejection data
            cull_df: DataFrame with cull data
            
        Returns:
            Dictionary mapping CSV type to file path
        """
        exported_files = {}
        
        # Export main harvest plan files
        exported_files['master'] = self.csv_exporter.export_harvest_plan_master(harvest_df)
        exported_files['slaughterhouse'] = self.csv_exporter.export_harvest_plan_slaughterhouse(harvest_df)
        exported_files['market'] = self.csv_exporter.export_harvest_plan_market(harvest_df)
        
        # Export cull data
        if cull_df is not None and not cull_df.empty:
            exported_files['culls'] = self.csv_exporter.export_harvest_plan_culls(cull_df)
        else:
            exported_files['culls'] = self.csv_exporter.export_harvest_plan_culls(pd.DataFrame())
        
        # Export summaries
        exported_files['summary_farms'] = self.csv_exporter.export_summary_farms(harvest_df)
        exported_files['summary_houses'] = self.csv_exporter.export_summary_houses(harvest_df)
        
        # Export unharvested stock
        if unharvested_df is not None:
            exported_files['unharvested_stock'] = self.csv_exporter.export_unharvested_stock(unharvested_df, ready_df, best_market_plan, best_sh_plan)
        else:
            exported_files['unharvested_stock'] = self.csv_exporter.export_unharvested_stock(pd.DataFrame(), ready_df, best_market_plan, best_sh_plan)
        
        # Export rejection reasons
        if rejection_df is not None:
            exported_files['rejection_reasons'] = self.csv_exporter.export_rejection_reasons(rejection_df)
        else:
            exported_files['rejection_reasons'] = self.csv_exporter.export_rejection_reasons(pd.DataFrame())

        # Export harvest progression
        if harvest_df_progress is not None:
            exported_files['harvest_progression'] = self.csv_exporter.export_harvest_progression(harvest_df_progress)
        else:
            exported_files['harvest_progression'] = self.csv_exporter.export_harvest_progression(pd.DataFrame())

        # Export daily summary
        exported_files['daily_summary'] = self.csv_exporter.export_daily_harvest_summary(harvest_df)
        
        logger.info(f"Successfully exported {len(exported_files)} CSV files")
        return exported_files